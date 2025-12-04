from selenium import webdriver
from selenium.webdriver.edge.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select, WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import time, os, csv, glob
import pandas as pd
import numpy as np

downloads_folder = os.path.join(os.path.expanduser("~"), "Downloads")
pattern = os.path.join(downloads_folder, "exchange-rates*.csv")

# Find all matching files
files = glob.glob(pattern)

if files:
    print(f"Found {len(files)} file(s). Deleting...")
    for f in files:
        try:
            os.remove(f)
            print(f"Deleted: {f}")
        except Exception as e:
            print(f"Error deleting {f}: {e}")
else:
    print("No matching files found.")

driver_path = "msedgedriver.exe"
os.system("taskkill /f /im msedge.exe")

options = webdriver.EdgeOptions()
options.use_chromium = True
options.add_argument("start-maximized")

service = Service(driver_path)
driver = webdriver.Edge(service=service, options=options)

url = "https://www.bnm.gov.my/exchange-rates"
driver.get(url)

time.sleep(10)
wait = WebDriverWait(driver, 20)

rate_types = ["Buying", "Middle Rate", "Selling"]

for rate in rate_types:
# ===== SELECT DROPDOWN VALUES =====
    Select(wait.until(EC.presence_of_element_located((By.ID, "_bnm_exchange_rate_display_portlet_monthStart")))).select_by_visible_text("November")
    Select(wait.until(EC.presence_of_element_located((By.ID,"_bnm_exchange_rate_display_portlet_rateType")))).select_by_visible_text(rate)
    Select(wait.until(EC.presence_of_element_located((By.ID,"_bnm_exchange_rate_display_portlet_quotation")))).select_by_visible_text("Ringgit/Foreign Currency")

    # Click Search
    wait.until(EC.element_to_be_clickable((By.ID,"_bnm_exchange_rate_display_portlet_btnSearch"))).click()
    time.sleep(10)  # allow results to load
    wait.until(EC.element_to_be_clickable((By.CLASS_NAME,"btn-primary"))).click()
    time.sleep(20)

driver.quit()

def split_csv_tables(file_path):
    df = pd.read_csv(file_path, header=None)
    df = df.map(lambda x: str(x).strip() if not pd.isna(x) else x)

    tables = []
    current_table = []

    for _, row in df.iterrows():
        if all((pd.isna(x) or x == "" or x == "Â") for x in row):
            if current_table:
                tbl = pd.DataFrame(current_table[1:], columns=current_table[0])
                tables.append(tbl)
                current_table = []
        else:
            current_table.append(row.tolist())

    if current_table:
        tbl = pd.DataFrame(current_table[1:], columns=current_table[0])
        tables.append(tbl)

    return tables

files = glob.glob(pattern)   # find all CSV files starting with "exchange-rates"
all_merged = []   # store each merged result (Buying/Middle/Selling)
name_map = {
    "(1)": "Middle Rate",
    "(2)": "Selling",
    "s": "Buying"
}
merged_by_filename = {} 
output_rows = []     # will store each row for the CSV
countries = []        # to list all country labels

# Create final column structure
final_df = pd.DataFrame(columns=["Country", "Middle Rate", "Selling", "Buying"])

for f in files:
    tables = split_csv_tables(f)

    print(f"Tables found: {len(tables)}")

    ### JOIN all tables by the first column
    for i in range(len(tables)):
        key_col = tables[i].columns[0]
        tables[i] = tables[i].rename(columns={ key_col : "Date" })

    from functools import reduce
    merged_df = reduce(lambda left, right: pd.merge(left, right, on="Date", how="outer"), tables)
    merged_df["Date"] = pd.to_datetime(merged_df["Date"], dayfirst=True)
    merged_df = merged_df.sort_values(by="Date", ascending=False)

    base = os.path.basename(f)
    name_no_ext = os.path.splitext(base)[0]
    mapped_name = name_no_ext
    for key, label in name_map.items():
        if key in name_no_ext:
            mapped_name = name_no_ext.replace(key, "-" + label)
            break

    merged_by_filename[mapped_name] = merged_df

final_dict = {}   # { "USD": {"Middle Rate": x, "Selling": y, "Buying": z}, ... }

for name, df in merged_by_filename.items():

    print(f"\n---- Processing: {name} ----")
    print(df)

    label = name.split("-")[-1].strip()   # Middle Rate / Selling / Buying

    # Skip empty df
    if df is None or df.empty:
        print("Empty dataframe, skipping:", name)
        continue

    # Take FIRST ROW only (latest date after sorting)
    row = df.iloc[0]

    # Loop through all currency columns (skip first col: Date)
    for currency in df.columns[1:]:
        
        rate_value = row[currency]

        # Ensure currency exists in dictionary
        if currency not in final_dict:
            final_dict[currency] = {"Middle Rate": "", "Selling": "", "Buying": ""}

        # Assign the rate to the right place
        final_dict[currency][label] = rate_value

# Convert final_dict → DataFrame
final_df = pd.DataFrame([
    {"Country": c, 
     "Middle Rate": vals["Middle Rate"],
     "Selling": vals["Selling"],
     "Buying": vals["Buying"]}
    for c, vals in final_dict.items()
])

# Sort alphabetically
final_df = final_df.sort_values("Country")

# Save CSV
output_path = r"compiled_rates.csv"
final_df.to_csv(output_path, index=False)

print("\nCSV CREATED SUCCESSFULLY:", output_path)
print("\nPreview:")
print(final_df)

file = r"excel_sample_data_qae.xlsx"

# Sheet names
sheet_sales = "python_test-sales"
sheet_product = "python_test-product"
sheet_store = "python_test-store"

# Read sheets into DataFrames
sales_df = pd.read_excel(file, sheet_name=sheet_sales)
product_df = pd.read_excel(file, sheet_name=sheet_product)
store_df = pd.read_excel(file, sheet_name=sheet_store)

sales_csv = r"sales.csv"
product_csv = r"product.csv"
store_csv = r"store.csv"

sales_df.to_csv(sales_csv, index=False)
product_df.to_csv(product_csv, index=False)
store_df.to_csv(store_csv, index=False)

print("\nCSV FILES CREATED SUCCESSFULLY:")
print(sales_df)
print(product_df)
print(store_df)

combine_df = sales_df.merge(product_df, on="product_code", how="left")
combine_df = combine_df.merge(store_df, on="store_code", how="left")
print(combine_df)

final_df["currency_key"] = final_df["Country"].str[:3].str.upper()
combine_df["currency"] = combine_df["currency"].str.upper()

complete_df = combine_df.merge(
    final_df,
    left_on="currency",
    right_on="currency_key",
    how="left"
)
cols_to_fill = ["Middle Rate", "Selling", "Buying"]
complete_df[cols_to_fill] = complete_df[cols_to_fill].fillna(1.0)
print(complete_df)

# Unique region and category choices
available_regions = sorted(complete_df["store_region"].dropna().unique())
available_categories = sorted(complete_df["product_category"].dropna().unique())
available_regions_lower = [r.lower() for r in available_regions]
available_categories_lower = [c.lower() for c in available_categories]

# -----------------------
# REGION SELECTION (LOOP)
# -----------------------
print("Available Regions:")
for r in available_regions:
    print(" -", r)

while True:
    region_input = input("\nEnter region (or 'ALL' to include all regions): ").strip()

    if region_input.lower() == "all":
        filtered_region = complete_df.copy()
        break

    if region_input.lower() in available_regions_lower:
        # Find the original cased value
        matched_region = available_regions[available_regions_lower.index(region_input.lower())]
        filtered_region = complete_df[complete_df["store_region"] == matched_region]
        break

    print("❌ Invalid region. Please choose from the list above.")

# -------------------------------
# PRODUCT CATEGORY SELECTION LOOP
# -------------------------------
print("\nAvailable Product Categories:")
for c in available_categories:
    print(" -", c)

while True:
    category_input = input("\nEnter product category (or 'ALL' to include all categories): ").strip()

    if category_input.lower() == "all":
        filtered_category = complete_df.copy()
        break

    if category_input.lower() in available_categories_lower:
        matched_category = available_categories[available_categories_lower.index(category_input.lower())]
        filtered_category = complete_df[complete_df["product_category"] == matched_category]
        break

    print("❌ Invalid category. Please choose from the list above.")

# Ensure Middle Rate is numeric
filtered_region["Buying"] = pd.to_numeric(filtered_region["Buying"], errors="coerce").round(2)
filtered_region["Selling"] = pd.to_numeric(filtered_region["Selling"], errors="coerce").round(2)
filtered_category["Buying"] = pd.to_numeric(filtered_category["Buying"], errors="coerce").round(2)
filtered_category["Selling"] = pd.to_numeric(filtered_category["Selling"], errors="coerce").round(2)

# Convert price & cost to MYR
filtered_region["price_myr"] = (filtered_region["price"] * filtered_region["Buying"]).round(2)
filtered_region["cost_myr"] = (filtered_region["cost"] * filtered_region["Selling"]).round(2)
filtered_category["price_myr"] = (filtered_category["price"] * filtered_category["Buying"]).round(2)
filtered_category["cost_myr"] = (filtered_category["cost"] * filtered_category["Selling"]).round(2)

# Sales amount & cost in MYR
filtered_region["sales_amount"] = filtered_region["sales_qty"] * filtered_region["price_myr"]
filtered_region["sales_cost"]   = filtered_region["sales_qty"] * filtered_region["cost_myr"]
filtered_category["sales_amount"] = filtered_category["sales_qty"] * filtered_category["price_myr"]
filtered_category["sales_cost"]   = filtered_category["sales_qty"] * filtered_category["cost_myr"]

# Profit
filtered_region["profit"] = filtered_region["sales_amount"] - filtered_region["sales_cost"]
filtered_category["profit"] = filtered_category["sales_amount"] - filtered_category["sales_cost"]

report_region = (
    filtered_region.groupby("store_region")
    .agg({
        "sales_qty": "sum",
        "sales_amount": "sum",
        "sales_cost": "sum",
        "profit": "sum"
    })
    .reset_index()
)

report_product = (
    filtered_category.groupby("product_category")
    .agg({
        "sales_qty": "sum",
        "sales_amount": "sum",
        "sales_cost": "sum",
        "profit": "sum"
    })
    .reset_index()
)

# Format to 2 decimal places (only for amount columns)
amount_cols = ["sales_amount", "sales_cost", "profit"]
report_region[amount_cols] = report_region[amount_cols].round(2)
report_product[amount_cols] = report_product[amount_cols].round(2)

print("\n===== SALES REPORT BY REGION (MYR)=====")
print(report_region)

print("\n===== SALES REPORT BY PRODUCT CATEGORY (MYR)=====")
print(report_product)

output_path = r"sales_report.xlsx"

with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
    report_region.to_excel(writer, sheet_name="By Region", index=False)
    report_product.to_excel(writer, sheet_name="By Product Category", index=False)

# ------------------------------------------------------
# APPLY FORMATTING
# ------------------------------------------------------

wb = load_workbook(output_path)

thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

def format_sheet(ws):
    # Bold header + center alignment
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Apply borders + auto width for all cells
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border

    # Auto column width
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

# Format both sheets
format_sheet(wb["By Region"])
format_sheet(wb["By Product Category"])

wb.save(output_path)

print("\nExcel report created successfully:")
print(output_path)

